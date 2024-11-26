import pandas as pd
import re
from openpyxl.styles import Border, Side, Alignment,PatternFill,Font
from tkinter import Tk
from tkinter.filedialog import askopenfilenames
from openpyxl import Workbook
from openpyxl.worksheet.page import PageMargins
import datetime


root = Tk()
root.withdraw()  # Tkinter 창을 숨김
file_paths = []
convert_file_paths = askopenfilenames(title="엑셀 파일 선택", filetypes=[("Excel 파일", "*.xls")])

    # 현재 디렉터리에서 .xls 확장자를 가진 파일 목록 찾기
    # 파일 목록 출력

for convert_file in convert_file_paths:
    try:
        # HTML 파일을 pandas로 읽기
        df = pd.read_html(convert_file, flavor='lxml')[0]  # 첫 번째 테이블 읽기
        
        # 파일명에서 확장자를 제거하고 .xlsx 확장자 추가
        new_file = convert_file.replace('.xls', 'covert.xlsx')
        
        # Excel 파일로 저장
        df.to_excel(new_file, index=False, engine='openpyxl')
        
        print(f"{convert_file} -> {new_file} 변환 완료")

        file_paths.append(new_file)

    except Exception as e:
        print(f"{convert_file} 변환 실패: {e}")
        


        
def Two_cell_merge(j):
    start_site = new_index_values.index(new_index_values[j])
    end_site = new_column_names.index(new_column_names[change_colums])
    merge.append((start_site,end_site))
     
def Two_cell_fill(j):
    index_value = new_index_values.index(new_index_values[j])
    column_name = new_column_names.index(new_column_names[change_colums])
    specific_cells.append((index_value, column_name))
    index_value = new_index_values.index(new_index_values[j+1])
    column_name = new_column_names.index(new_column_names[change_colums])
    specific_cells.append((index_value, column_name))

def Four_cell_fill(j_1):
    index_value = new_index_values.index(new_index_values[j_1])
    column_name = new_column_names.index(new_column_names[change_colums])
    specific_cells.append((index_value,column_name))
    index_value = new_index_values.index(new_index_values[j_1+1])
    column_name = new_column_names.index(new_column_names[change_colums])
    specific_cells.append((index_value, column_name))
    index_value = new_index_values.index(new_index_values[j_1+2])
    column_name = new_column_names.index(new_column_names[change_colums])
    specific_cells.append((index_value, column_name))
    index_value = new_index_values.index(new_index_values[j_1+3])
    column_name = new_column_names.index(new_column_names[change_colums])
    specific_cells.append((index_value, column_name))

def Four_cell_merge(j_1):
    start_site = new_index_values.index(new_index_values[j_1])
    end_site = new_column_names.index(new_column_names[change_colums])
    merges.append((start_site,end_site))     

#현장결제 버그 수정됨

print("2024.11.26일 버전")
print("하계 버전")
 # Tkinter 창을 숨김

# 파일 선택 다이얼로그를 통해 파일 경로 획득

for file_path in file_paths:

    
    # DataFrame으로 읽기
    date=['월요일','화요일','수요일','목요일','금요일','토요일','일요일']
    soft_tennis_value= ['안성맞춤소프트테니스구장(테니스구장(1코트))','안성맞춤소프트테니스구장(테니스구장(2코트))','안성맞춤소프트테니스구장(테니스구장(3코트))','안성맞춤소프트테니스구장(테니스구장(4코트))','안성맞춤소프트테니스구장(테니스구장(5코트))','안성맞춤소프트테니스구장(테니스구장(6코트))','안성맞춤소프트테니스구장(테니스구장(7코트))','안성맞춤소프트테니스구장(테니스구장(8코트))']
    tennis_value = ['안성맞춤테니스구장(테니스구장(9코트))','안성맞춤테니스구장(테니스구장(10코트))','안성맞춤테니스구장(테니스구장(11코트))','안성맞춤테니스구장(테니스구장(12코트))']
    df_data = pd.read_excel(file_path, index_col=0)

    #셀에 있는 날짜를 파이썬이 읽고 변환
    date_string = df_data['예약일'][1].replace('.','-')
    date_object = datetime.datetime.strptime(date_string, '%Y-%m-%d')

    # weekday 메서드와 변환된 data_object를 사용하여 요일을 숫자로 얻기 (0: 월요일, 1: 화요일, ..., 6: 일요일)
    day_of_week_number = date_object.weekday()

    #셀에 있는 시설명이 테니스 벨류면 실행함    
    if df_data['시설명'].isin(tennis_value).any():
        new_column_names = ['9코트', '10코트', '11코트','12코트','기타'] # 필요한 만큼 열 이름을 변경
        new_index_values = ['06:00~07:00', '07:00~08:00', '08:00~09:00', '09:00~10:00', '10:00~11:00', '11:00~12:00', '12:00~13:00','13:00~14:00','14:00~15:00','15:00~16:00','16:00~17:00','17:00~18:00','18:00~19:00','19:00~20:00','20:00~21:00','21:00~22:00'] #행
        specific_cells = []
        merge = []
        merges =[]
        notsign_text = "❨미인증❩\n\n 사용자      ❨서명❩ \n\n관리자      ❨서명❩"
        notpaidsign_text = "❨현장결제❩\n\n 사용자      ❨서명❩ \n\n관리자      ❨서명❩"
        sign_text = "\n\n 사용자      ❨서명❩ \n\n관리자      ❨서명❩"
        other_contry = "❨관외❩\n\n 관외사용자    ❨서명❩ \n\n관리자      ❨서명❩"


        #조건 리스트
        desired_reservation_status_list = ['결제가능', '상담대기', '예약완료']
        desired_reservation_notpaid = ['현장결제']
        desired_facility_list= ['안성맞춤테니스구장(테니스구장(9코트))','안성맞춤테니스구장(테니스구장(10코트))','안성맞춤테니스구장(테니스구장(11코트))','안성맞춤테니스구장(테니스구장(12코트))']
        desired_reservation_time_list = ['06:00~08:00', '08:00~10:00', '10:00~12:00', '12:00~14:00', '14:00~16:00', '16:00~18:00', '18:00~20:00', '20:00~22:00']
        desired_reservation_time_list_4 = ['06:00~10:00', '08:00~12:00', '10:00~14:00', '12:00~16:00', '14:00~18:00', '16:00~20:00', '18:00~22:00']
        desired_reservation_time_list_r = ['6-8','8-10','10-12','12-14','14-16','16-18','18-20','20-22']
        desired_reservation_time_list_4_r = ['6-10', '8-12', '10-14', '12-16', '14-18', '16-20', '18-22']
        
        #라이트 사용 여부 판별
        desired_money = [3000,6000]
        desired_money_zero=[0,3000,6000]
        
        # 미인증 안성시민 판별
        no_certifiacte_people_week = [10000,13000,20000,26000]
        no_certifiacte_people_weekend = [13000, 18000,26000,36000]
        
        # 새로운 엑셀 파일을 생성
        df_sch = pd.DataFrame(index=new_index_values, columns=new_column_names)

        # 9~12번 코트
        for change_colums in range(4):
            j=0
            j_1=0
        

            for i in range (8): # desired_reservation_time_list 속의 value 값의 수

                #현장결제
                if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                   condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list[i])& (df_data['예약상태'].isin(desired_reservation_notpaid))
                   
                   if condition.any():
                        reserved_member = condition[condition].index[0]
                        combined_value = f"{reserved_member} {desired_reservation_time_list_r[i]} {notpaidsign_text}" # 엑셀에 쓰여질 문구
                        df_sch.loc[[new_index_values[j],new_index_values[j+1]], new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치

                        Two_cell_merge (j)


                # 라이트 수동 추가
                if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                    condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list[i])& (df_data['예약상태'].isin(desired_reservation_status_list))&(df_data['추가금액'].isin(desired_money))

                    if condition.any():
                        reserved_member = condition[condition].index[0]
                        combined_value = f"{reserved_member} {desired_reservation_time_list_r[i]} {sign_text}" # 엑셀에 쓰여질 문구
                        df_sch.loc[[new_index_values[j],new_index_values[j+1]], new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치

                        Two_cell_merge (j)

                        Two_cell_fill(j)
                        

                #관외 라이트
                if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                    condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list[i])& (df_data['예약상태'].isin(desired_reservation_status_list))&(abs(df_data['할인전금액']-df_data['추가금액'])==3000)

                    if condition.any():
                        # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                        reserved_member = condition[condition].index[0]
                        combined_value = f"{reserved_member} {desired_reservation_time_list_r[i]} {other_contry}"# 엑셀에 쓰여질 문구
                        df_sch.loc[[new_index_values[j],new_index_values[j+1]], new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치

                        Two_cell_merge (j)

                        Two_cell_fill(j)      
                        

                #라이트 자동 추가   
                if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                    condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list[i])& (df_data['예약상태'].isin(desired_reservation_status_list))&(~df_data['추가금액'].isin(desired_money))&(df_data['할인전금액'] - (df_data['할인금액'] * 5 / 4) == 3000)
                                        
                    if condition.any():
                        # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                        reserved_member = condition[condition].index[0]
                        combined_value = f"{reserved_member} {desired_reservation_time_list_r[i]} {sign_text}" # 엑셀에 쓰여질 문구
                        df_sch.loc[[new_index_values[j],new_index_values[j+1]], new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치

                        Two_cell_merge (j)

                        #셀 색칠
                        Two_cell_fill(j)
                
                #일반 예약

                    if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                        condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list[i])& (df_data['예약상태'].isin(desired_reservation_status_list))&(df_data['추가금액']==0)

                        if condition.any():
                        # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                            reserved_member = condition[condition].index[0]
                            combined_value = f"{reserved_member} {desired_reservation_time_list_r[i]} {sign_text}"# 엑셀에 쓰여질 문구
                            df_sch.loc[[new_index_values[j],new_index_values[j+1]], new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치

                            Two_cell_merge (j)
                



                #일반 관외 
                if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                    condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list[i])& (df_data['예약상태'].isin(desired_reservation_status_list))&(~df_data['추가금액'].isin(desired_money_zero))

                    if condition.any():
                        # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                        reserved_member = condition[condition].index[0]
                        combined_value = f"{reserved_member} {desired_reservation_time_list_r[i]} {other_contry}"# 엑셀에 쓰여질 문구
                        df_sch.loc[[new_index_values[j],new_index_values[j+1]], new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치

                        Two_cell_merge (j)


    # 미인증 시민 라이트

                # 주말일 경우 금액
                if day_of_week_number == 6 or day_of_week_number == 5:
                    
                    if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                        condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list[i])& (df_data['예약상태'].isin(desired_reservation_status_list))&(df_data['추가금액']== 0)&(df_data['할인금액']== 0 )&(df_data['결제금액'].isin(no_certifiacte_people_weekend))
                        if condition.any():
                                reserved_member = condition[condition].index[0]
                                combined_value = f"{reserved_member} {desired_reservation_time_list_r[i]} {notsign_text}"# 엑셀에 쓰여질 문구
                                df_sch.loc[[new_index_values[j],new_index_values[j+1]], new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치

                                Two_cell_merge (j)
                              #셀 색칠
                                Two_cell_fill(j)

                # 평일일 경우 금액
                if day_of_week_number != 6 and day_of_week_number != 5:
                   
                    if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                        condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list[i])& (df_data['예약상태'].isin(desired_reservation_status_list))&(df_data['추가금액']==0)&(df_data['할인금액']==0)&(df_data['결제금액'].isin(no_certifiacte_people_week))
                        if condition.any():
                            reserved_member = condition[condition].index[0]
                            combined_value = f"{reserved_member} {desired_reservation_time_list_r[i]} {notsign_text}"# 엑셀에 쓰여질 문구
                            df_sch.loc[[new_index_values[j],new_index_values[j+1]], new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치

                            Two_cell_merge (j)

                            Two_cell_fill(j)

                

                j=j+2
                

            for k in range(7):
                    #현장결제
                if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                   condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list_4[k])& (df_data['예약상태'].isin(desired_reservation_notpaid))
                   
                   if condition.any():
                        reserved_member = condition[condition].index[0]
                        combined_value = f"{reserved_member} {desired_reservation_time_list_4_r[k]} {notpaidsign_text}" # 엑셀에 쓰여질 문구
                        df_sch.loc[[new_index_values[j_1],new_index_values[j_1+1],new_index_values[j_1+2],new_index_values[j_1+3]],new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치

                        Four_cell_merge(j_1)


                    #라이트 수동추가 
                if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                        condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list_4[k])& (df_data['예약상태'].isin(desired_reservation_status_list))&(df_data['추가금액'].isin(desired_money))
                        
                        if condition.any():
                            # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                            reserved_member = condition[condition].index[0]
                            combined_value = f"{reserved_member} {desired_reservation_time_list_4_r[k]} {sign_text}"# 엑셀에 쓰여질 문구
                            df_sch.loc[[new_index_values[j_1],new_index_values[j_1+1],new_index_values[j_1+2],new_index_values[j_1+3]] , new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치
                            
                            Four_cell_merge(j_1)
                            #셀 색칠
                            Four_cell_fill(j_1)
                            


                    #관외 라이트 추가
                if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                        condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list_4[k])& (df_data['예약상태'].isin(desired_reservation_status_list))&(abs(df_data['할인전금액']-df_data['추가금액'])==6000)
                        if condition.any():
                            # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                            reserved_member = condition[condition].index[0]
                            combined_value = f"{reserved_member} {desired_reservation_time_list_4_r[k]}{other_contry}"# 엑셀에 쓰여질 문구
                            df_sch.loc[[new_index_values[j_1],new_index_values[j_1+1],new_index_values[j_1+2],new_index_values[j_1+3]] , new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치
                            
                            
                            
                            Four_cell_merge(j_1)
                            
                            
                            #셀 색칠
                            Four_cell_fill(j_1)


                    #일반 라이트 자동추가
                if(df_data['시설명'] == desired_facility_list[change_colums]).any():
                        condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list_4[k])& (df_data['예약상태'].isin(desired_reservation_status_list))&(~df_data['추가금액'].isin(desired_money))&(df_data['할인전금액'] - (df_data['할인금액'] * 5 /4 ) == 6000)


                        if condition.any():
                            # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                            reserved_member = condition[condition].index[0]
                            combined_value = f"{reserved_member} {desired_reservation_time_list_4_r[k]} {sign_text}"# 엑셀에 쓰여질 문구
                            df_sch.loc[[new_index_values[j_1],new_index_values[j_1+1],new_index_values[j_1+2],new_index_values[j_1+3]] , new_column_names[change_colums]] = combined_value
                                                        
                            Four_cell_merge(j_1)

                            #셀 색칠
                            Four_cell_fill(j_1)

                      # 일반 이용     
                if(df_data['시설명'] == desired_facility_list[change_colums]).any():
                        condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list_4[k])& (df_data['예약상태'].isin(desired_reservation_status_list))&(df_data['추가금액']==0)    

                        if condition.any():
                            # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                            reserved_member = condition[condition].index[0]
                            combined_value = f"{reserved_member} {desired_reservation_time_list_4_r[k]} {sign_text}"# 엑셀에 쓰여질 문구
                            df_sch.loc[[new_index_values[j_1],new_index_values[j_1+1],new_index_values[j_1+2],new_index_values[j_1+3]] , new_column_names[change_colums]] = combined_value     #엑셀에서 사용할 셀의 위치
                            
                            Four_cell_merge(j_1)

                    
                    #관외 일반이용
                if(df_data['시설명'] == desired_facility_list[change_colums]).any():
                        condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list_4[k])& (df_data['예약상태'].isin(desired_reservation_status_list))&(~df_data['추가금액'].isin(desired_money_zero))    

                        if condition.any():
                            # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                            reserved_member = condition[condition].index[0]
                            combined_value = f"{reserved_member} {desired_reservation_time_list_4_r[k]}{other_contry}"# 엑셀에 쓰여질 문구
                            df_sch.loc[[new_index_values[j_1],new_index_values[j_1+1],new_index_values[j_1+2],new_index_values[j_1+3]] , new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치    
                            
                            Four_cell_merge(j_1)
            
                    #미시민인증 주말
                if day_of_week_number == 6 or day_of_week_number == 5:
                       
                        if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                            condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list_4[k])& (df_data['예약상태'].isin(desired_reservation_status_list))&(df_data['추가금액']==0)&(df_data['할인금액']==0)&(df_data['결제금액'].isin(no_certifiacte_people_weekend))
                            
                            if condition.any():
                                reserved_member = condition[condition].index[0]
                                combined_value = f"{reserved_member} {desired_reservation_time_list_4_r[k]} {notsign_text}"# 엑셀에 쓰여질 문구
                                df_sch.loc[[new_index_values[j_1],new_index_values[j_1+1],new_index_values[j_1+2],new_index_values[j_1+3]] , new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치
                                       
                                Four_cell_merge(j_1)
                                    
                                #셀 색칠
                                Four_cell_fill(j_1)
                                    
                    #미시민인증 평일
                if day_of_week_number != 6 and day_of_week_number != 5:
                        
                        if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                            condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list_4[k])& (df_data['예약상태'].isin(desired_reservation_status_list))&(df_data['추가금액']==0)&(df_data['할인금액']==0)&(df_data['결제금액'].isin(no_certifiacte_people_week))
                            if condition.any():
                                eserved_member = condition[condition].index[0]
                                combined_value = f"{reserved_member} {desired_reservation_time_list_4_r[k]} {notsign_text}"# 엑셀에 쓰여질 문구
                                df_sch.loc[[new_index_values[j_1],new_index_values[j_1+1],new_index_values[j_1+2],new_index_values[j_1+3]] , new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치
                                    
                                Four_cell_merge(j_1)
                                    
                                    #셀 색칠
                                Four_cell_fill(j_1)
                j_1=j_1+2
            

        day_of_week_number = date[day_of_week_number]
        
        def remove_parentheses(value):
            return re.sub(r'\([^)]*\)', '', str(value))
        
        df_sch =df_sch.applymap(remove_parentheses)

        df_sch = df_sch.replace('nan', '')

        # ExcelWriter 객체 생성
        with pd.ExcelWriter(f" 하드 코트 {df_data['예약일'][1]}.xlsx", engine='openpyxl') as writer:
            # DataFrame을 Excel 파일에 쓰기
            df_sch.to_excel(writer, sheet_name='Sheet1', startcol=0, startrow=1, header=True, index=True)

            # ExcelWriter 객체에서 워크북과 워크시트 객체 가져오기
            workbook  = writer.book
            worksheet = writer.sheets['Sheet1']

            red_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            for position in specific_cells:
                row, col = position
                cell = worksheet.cell(row=row + 3, column=col + 2)  # 엑셀은 1부터 시작하므로 +1
                cell.fill = red_fill


            #병합하기
            for positions in merge:
                row, col = positions
                worksheet.merge_cells(start_row=row + 3, start_column=col + 2, end_row=row + 4, end_column=col + 2)

            for positions in merges:
                row, col = positions
                worksheet.merge_cells(start_row=row + 3, start_column=col + 2, end_row=row + 6, end_column=col + 2)





            # 열의 너비를 15로 설정
            for col_num, value in enumerate(new_column_names):
                worksheet.column_dimensions[worksheet.cell(row=2, column=col_num+1).column_letter].width = 25

            # 높이 설정
            for row_num, value in enumerate(new_index_values):
                worksheet.row_dimensions[row_num + 3].height = 33

            # 헤더 텍스트 추가
            header_text = f"                            테니스장 (하드 코트)          {df_data['예약일'][1]} {day_of_week_number}"  # 페이지 번호를 나타내는 예시
            worksheet['A1'] = header_text
            worksheet.merge_cells('A1:E1')

            # A1 셀의 높이를 늘리기
            worksheet.row_dimensions[1].height = 30

            worksheet['A1'].font = Font(size=16)
            # 외곽에 선 추가
            border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

            # 안쪽에 선 추가
            inside_border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin', color='000000'))

            # 셀에 스타일 적용
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(size=9)

            # 안쪽에 선 추가
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = inside_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(size=9)

            worksheet['A1'].font = Font(size=16, bold=True)

            worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE #가로로
            worksheet.page_setup.fitToPage = True #한페이지안에 넣기
            worksheet.page_setup.fitToHeight = 1 #배율 
            worksheet.page_margins = PageMargins(top=0, bottom=0, left=0, right=0) #프린트 여백
            
            print("생성 완료")

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    #  소프트테니스

    if df_data['시설명'].isin(soft_tennis_value).any(): 


        date=['월요일','화요일','수요일','목요일','금요일','토요일','일요일']
        new_column_names = ['1코트', '2코트', '3코트','4코트','5코트','6코트','7코트','8코트']  # 필요한 만큼 열 이름을 변경
        new_index_values = ['06:00~07:00', '07:00~08:00', '08:00~09:00', '09:00~10:00', '10:00~11:00', '11:00~12:00', '12:00~13:00','13:00~14:00','14:00~15:00','15:00~16:00','16:00~17:00','17:00~18:00','18:00~19:00','19:00~20:00','20:00~21:00','21:00~22:00'] #행
        specific_cells = []
        merge = []
        merges =[]
        notsign_text = "❨미인증❩\n\n 사용자      ❨서명❩ \n\n관리자      ❨서명❩"
        notpaidsign_text = "❨현장결제❩\n\n 사용자      ❨서명❩ \n\n관리자      ❨서명❩"
        sign_text = "\n\n 사용자      ❨서명❩ \n\n관리자      ❨서명❩"
        other_contry = " ❨관외❩\n\n 관외사용자    ❨서명❩ \n\n관리자      ❨서명❩"


        notsign_text_short_ver = "❨미인증❩\n 사용자      ❨서명❩ \n\n관리자      ❨서명❩"
        notpaidsign_text_short_ver = "❨현장결제❩\n 사용자      ❨서명❩ \n\n관리자      ❨서명❩"
        sign_text_short_ver = "\n사용자      ❨서명❩ \n\n관리자      ❨서명❩"
        other_contry_short_ver = "❨관외❩\n 관외사용자    ❨서명❩ \n\n관리자      ❨서명❩"

        #조건 리스트
        desired_reservation_status_list = ['결제가능', '상담대기', '예약완료']
        desired_reservation_notpaid = ['현장결제']
        desired_facility_list= ['안성맞춤소프트테니스구장(테니스구장(1코트))','안성맞춤소프트테니스구장(테니스구장(2코트))','안성맞춤소프트테니스구장(테니스구장(3코트))','안성맞춤소프트테니스구장(테니스구장(4코트))','안성맞춤소프트테니스구장(테니스구장(5코트))','안성맞춤소프트테니스구장(테니스구장(6코트))','안성맞춤소프트테니스구장(테니스구장(7코트))','안성맞춤소프트테니스구장(테니스구장(8코트))']
        desired_reservation_time_list = ['06:00~08:00', '08:00~10:00', '10:00~12:00', '12:00~14:00', '14:00~16:00', '16:00~18:00', '18:00~20:00', '20:00~22:00']
        desired_reservation_time_list_4 = ['06:00~10:00', '08:00~12:00', '10:00~14:00', '12:00~16:00', '14:00~18:00', '16:00~20:00', '18:00~22:00']
        desired_reservation_time_list_r = ['6-8','8-10','10-12','12-14','14-16','16-18','18-20','20-22']
        desired_reservation_time_list_4_r = ['6-10', '8-12', '10-14', '12-16', '14-18', '16-20', '18-22']

        desired_money = [4500,9000]
        desired_money_zero = [0,4500,9000]

        no_certifiacte_people_week =[8000,9500,16000,19000]
        no_certifiacte_people_weekend = [9500,12000,19000,24000]

        # 새로운 엑셀 파일을 생성
        df_sch = pd.DataFrame(index=new_index_values, columns=new_column_names)
        date_string = df_data['예약일'][1].replace('.','-') 
        date_object = datetime.datetime.strptime(date_string, '%Y-%m-%d')

        # weekday 메서드를 사용하여 요일을 숫자로 얻기 (0: 월요일, 1: 화요일, ..., 6: 일요일)
        day_of_week_number = date_object.weekday()

        #조건문 
        for change_colums in range(8):
            j=0
            j_1=0
            
            for i in range (8): # desired_reservation_time_list 속의 value 값의 수

                #현장결제
                if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                   condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list[i])& (df_data['예약상태'].isin(desired_reservation_notpaid))
                   
                   if condition.any():
                        reserved_member = condition[condition].index[0]
                        combined_value = f"{reserved_member} {desired_reservation_time_list_r[i]} {notpaidsign_text_short_ver}" # 엑셀에 쓰여질 문구
                        df_sch.loc[[new_index_values[j],new_index_values[j+1]], new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치

                        Two_cell_merge(j)


                # 라이트 수동 추가
                if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                    condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list[i])& (df_data['예약상태'].isin(desired_reservation_status_list))&(df_data['추가금액'].isin(desired_money))

                    if condition.any():
                        reserved_member = condition[condition].index[0]
                        combined_value = f"{reserved_member} {desired_reservation_time_list_r[i]} {sign_text_short_ver}" # 엑셀에 쓰여질 문구
                        df_sch.loc[[new_index_values[j],new_index_values[j+1]], new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치

                        Two_cell_merge(j)
                        
                        #셀 색칠
                        Two_cell_fill(j)

                #관외 라이트
                if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                    condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list[i])& (df_data['예약상태'].isin(desired_reservation_status_list))&(abs(df_data['할인전금액']-df_data['추가금액'])==4500)

                    if condition.any():
                        # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                        reserved_member = condition[condition].index[0]
                        combined_value = f"{reserved_member} {desired_reservation_time_list_r[i]} {other_contry_short_ver}"# 엑셀에 쓰여질 문구
                        df_sch.loc[[new_index_values[j],new_index_values[j+1]], new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치

                        Two_cell_merge(j)

                        #셀 색칠
                        Two_cell_fill(j)       
                        

                #라이트 자동 추가   
                if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                    condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list[i])& (df_data['예약상태'].isin(desired_reservation_status_list))&(~df_data['추가금액'].isin(desired_money))&(df_data['할인전금액'] - (df_data['할인금액'] * 5 / 4) == 4500)
                                        
                    if condition.any():
                        # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                        reserved_member = condition[condition].index[0]
                        combined_value = f"{reserved_member} {desired_reservation_time_list_r[i]} {sign_text_short_ver}" # 엑셀에 쓰여질 문구
                        df_sch.loc[[new_index_values[j],new_index_values[j+1]], new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치


                        Two_cell_merge(j)

                        #셀 색칠
                        Two_cell_fill(j)
                
                #일반 예약

                    if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                        condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list[i])& (df_data['예약상태'].isin(desired_reservation_status_list))&(df_data['추가금액']==0)

                        if condition.any():
                        # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                            reserved_member = condition[condition].index[0]
                            combined_value = f"{reserved_member} {desired_reservation_time_list_r[i]} {sign_text_short_ver}"# 엑셀에 쓰여질 문구
                            df_sch.loc[[new_index_values[j],new_index_values[j+1]], new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치


                            Two_cell_merge(j)
                
                #일반 관외 
                if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                    condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list[i])& (df_data['예약상태'].isin(desired_reservation_status_list))&(~df_data['추가금액'].isin(desired_money_zero))

                    if condition.any():
                        # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                        reserved_member = condition[condition].index[0]
                        combined_value = f"{reserved_member} {desired_reservation_time_list_r[i]} {other_contry_short_ver}"# 엑셀에 쓰여질 문구
                        df_sch.loc[[new_index_values[j],new_index_values[j+1]], new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치

                        Two_cell_merge(j)

    # 미인증 시민 라이트

                # 주말일 경우 금액
                if day_of_week_number == 6 or day_of_week_number == 5:
                    
                    if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                        condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list[i])& (df_data['예약상태'].isin(desired_reservation_status_list))&(df_data['추가금액']== 0)&(df_data['할인금액']== 0 )&(df_data['결제금액'].isin(no_certifiacte_people_weekend))
                        if condition.any():
                                reserved_member = condition[condition].index[0]
                                combined_value = f"{reserved_member} {desired_reservation_time_list_r[i]} {notsign_text_short_ver}"# 엑셀에 쓰여질 문구
                                df_sch.loc[[new_index_values[j],new_index_values[j+1]], new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치

                                Two_cell_merge(j)

                              #셀 색칠
                                Two_cell_fill(j)

                # 평일일 경우 금액
                if day_of_week_number != 6 and day_of_week_number != 5:
                   
                    if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                        condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list[i])& (df_data['예약상태'].isin(desired_reservation_status_list))&(df_data['추가금액']==0)&(df_data['할인금액']==0)&(df_data['결제금액'].isin(no_certifiacte_people_week))
                        if condition.any():
                            reserved_member = condition[condition].index[0]
                            combined_value = f"{reserved_member} {desired_reservation_time_list_r[i]} {notsign_text_short_ver}"# 엑셀에 쓰여질 문구
                            df_sch.loc[[new_index_values[j],new_index_values[j+1]], new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치

                            Two_cell_merge(j)

                            Two_cell_fill(j)

                j=j+2
                
            for k in range(7):
                    #현장결제
                if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                   condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list_4[k])& (df_data['예약상태'].isin(desired_reservation_notpaid))
                   
                   if condition.any():
                        reserved_member = condition[condition].index[0]
                        combined_value = f"{reserved_member} {desired_reservation_time_list_4_r[k]} {notpaidsign_text}" # 엑셀에 쓰여질 문구
                        df_sch.loc[[new_index_values[j_1],new_index_values[j_1+1],new_index_values[j_1+2],new_index_values[j_1+3]], new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치

                        Four_cell_merge(j_1)
                        
                    #라이트 수동추가 
                if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                        condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list_4[k])& (df_data['예약상태'].isin(desired_reservation_status_list))&(df_data['추가금액'].isin(desired_money))
                        
                        if condition.any():
                            # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                            reserved_member = condition[condition].index[0]
                            combined_value = f"{reserved_member} {desired_reservation_time_list_4_r[k]} {sign_text}"# 엑셀에 쓰여질 문구
                            df_sch.loc[[new_index_values[j_1],new_index_values[j_1+1],new_index_values[j_1+2],new_index_values[j_1+3]] , new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치
                            
                            Four_cell_merge(j_1)                            
                            #셀 색칠
                            Four_cell_fill(j_1)
                            
                    #관외 라이트 추가
                if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                        condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list_4[k])& (df_data['예약상태'].isin(desired_reservation_status_list))&(abs(df_data['할인전금액']-df_data['추가금액'])==9000)
                        if condition.any():
                            # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                            reserved_member = condition[condition].index[0]
                            combined_value = f"{reserved_member} {desired_reservation_time_list_4_r[k]}{other_contry}"# 엑셀에 쓰여질 문구
                            df_sch.loc[[new_index_values[j_1],new_index_values[j_1+1],new_index_values[j_1+2],new_index_values[j_1+3]] , new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치
                            
                            Four_cell_merge(j_1)
                            #셀 색칠
                            Four_cell_fill(j_1)


                    #일반 라이트 자동추가
                if(df_data['시설명'] == desired_facility_list[change_colums]).any():
                        condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list_4[k])& (df_data['예약상태'].isin(desired_reservation_status_list))&(~df_data['추가금액'].isin(desired_money))&(df_data['할인전금액'] - (df_data['할인금액'] * 5 /4 ) == 9000)

                        if condition.any():
                            # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                            reserved_member = condition[condition].index[0]
                            combined_value = f"{reserved_member} {desired_reservation_time_list_4_r[k]} {sign_text}"# 엑셀에 쓰여질 문구
                            df_sch.loc[[new_index_values[j_1],new_index_values[j_1+1],new_index_values[j_1+2],new_index_values[j_1+3]] , new_column_names[change_colums]] = combined_value
                            
                            Four_cell_merge(j_1)
                            
                            #셀 색칠
                            Four_cell_fill(j_1)

                      # 일반 이용     
                if(df_data['시설명'] == desired_facility_list[change_colums]).any():
                        condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list_4[k])& (df_data['예약상태'].isin(desired_reservation_status_list))&(df_data['추가금액']==0)    

                        if condition.any():
                            # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                            reserved_member = condition[condition].index[0]
                            combined_value = f"{reserved_member} {desired_reservation_time_list_4_r[k]} {sign_text}"# 엑셀에 쓰여질 문구
                            df_sch.loc[[new_index_values[j_1],new_index_values[j_1+1],new_index_values[j_1+2],new_index_values[j_1+3]] , new_column_names[change_colums]] = combined_value     #엑셀에서 사용할 셀의 위치
                            
                            Four_cell_merge(j_1)

                    #관외 일반이용
                if(df_data['시설명'] == desired_facility_list[change_colums]).any():
                        condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list_4[k])& (df_data['예약상태'].isin(desired_reservation_status_list))&(~df_data['추가금액'].isin(desired_money_zero))    

                        if condition.any():
                            # 조건을 만족하면 해당 행의 인덱스인 '예약회원'을 출력
                            reserved_member = condition[condition].index[0]
                            combined_value = f"{reserved_member} {desired_reservation_time_list_4_r[k]}{other_contry}"# 엑셀에 쓰여질 문구
                            df_sch.loc[[new_index_values[j_1],new_index_values[j_1+1],new_index_values[j_1+2],new_index_values[j_1+3]] , new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치    
                            
                            Four_cell_merge(j_1)
            
                    #미시민인증 주말
                if day_of_week_number == 6 or day_of_week_number == 5:
                       
                        if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                            condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list_4[k])& (df_data['예약상태'].isin(desired_reservation_status_list))&(df_data['추가금액']==0)&(df_data['할인금액']==0)&(df_data['결제금액'].isin(no_certifiacte_people_weekend))
                            
                            if condition.any():
                                reserved_member = condition[condition].index[0]
                                combined_value = f"{reserved_member} {desired_reservation_time_list_4_r[k]} {notsign_text}"# 엑셀에 쓰여질 문구
                                df_sch.loc[[new_index_values[j_1],new_index_values[j_1+1],new_index_values[j_1+2],new_index_values[j_1+3]] , new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치
                                
                                Four_cell_merge(j_1)
                                    
                                #셀 색칠
                                Four_cell_fill(j_1)
                                    
                    #미시민인증 평일
                if day_of_week_number != 6 and day_of_week_number != 5:
                        
                        if (df_data['시설명'] == desired_facility_list[change_colums]).any():
                            condition = (df_data['시설명'] == desired_facility_list[change_colums]) & (df_data['예약시간'] == desired_reservation_time_list_4[k])& (df_data['예약상태'].isin(desired_reservation_status_list))&(df_data['추가금액']==0)&(df_data['할인금액']==0)&(df_data['결제금액'].isin(no_certifiacte_people_week))
                            if condition.any():
                                eserved_member = condition[condition].index[0]
                                combined_value = f"{reserved_member} {desired_reservation_time_list_4_r[k]} {notsign_text}"# 엑셀에 쓰여질 문구
                                df_sch.loc[[new_index_values[j_1],new_index_values[j_1+1],new_index_values[j_1+2],new_index_values[j_1+3]] , new_column_names[change_colums]] = combined_value #엑셀에서 사용할 셀의 위치
                                    
                                Four_cell_merge(j_1)
                                
                                    #셀 색칠
                                Four_cell_fill(j_1)
                j_1=j_1+2
        
        # 날짜 문자열을 datetime 객체로 변환
        date_string = df_data['예약일'][1].replace('.','-') 
        date_object = datetime.datetime.strptime(date_string, '%Y-%m-%d')

        # weekday 메서드를 사용하여 요일을 숫자로 얻기 (0: 월요일, 1: 화요일, ..., 6: 일요일)
        day_of_week_number = date_object.weekday()

        day_of_week_number = date[day_of_week_number]

        def remove_parentheses(value):
            return re.sub(r'\([^)]*\)', '', str(value))

        df_sch = df_sch.applymap(remove_parentheses)
        df_sch = df_sch.replace('nan', '')

        # ExcelWriter 객체 생성
        with pd.ExcelWriter(f" 정구장 {df_data['예약일'][1]}.xlsx", engine='openpyxl') as writer:
            # DataFrame을 Excel 파일에 쓰기
            df_sch.to_excel(writer, sheet_name='Sheet1', startcol=0, startrow=1, header=True, index=True)

            # ExcelWriter 객체에서 워크북과 워크시트 객체 가져오기
            workbook  = writer.book
            worksheet = writer.sheets['Sheet1']

            red_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            for position in specific_cells:
                row, col = position
                cell = worksheet.cell(row=row + 3, column=col + 2)  # 엑셀은 1부터 시작하므로 +1
                cell.fill = red_fill

            #병합하기이이이이이이
            
            for positions in merge:
                row, col = positions
                worksheet.merge_cells(start_row=row + 3, start_column=col + 2, end_row=row + 4, end_column=col + 2)

            for positions in merges:
                row, col = positions
                worksheet.merge_cells(start_row=row + 3, start_column=col + 2, end_row=row + 6, end_column=col + 2)

            # 열의 너비를 15로 설정
            for col_num, value in enumerate(new_column_names):
                worksheet.column_dimensions[worksheet.cell(row=2, column=col_num+1).column_letter].width = 15

            # 높이 설정
            for row_num, value in enumerate(new_index_values):
                worksheet.row_dimensions[row_num + 3].height = 33

            # 헤더 텍스트 추가
            header_text = f"                            소프트테니스구장             {df_data['예약일'][1]} {day_of_week_number}"  # 페이지 번호를 나타내는 예시
            worksheet['A1'] = header_text
            worksheet.merge_cells('A1:I1')

            # A1 셀의 높이를 늘리기
            worksheet.row_dimensions[1].height = 30

            worksheet['A1'].font = Font(size=16)
            # 외곽에 선 추가
            border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

            # 안쪽에 선 추가
            inside_border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin', color='000000'))
            
            # 셀에 스타일 적용
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(size=9)

            # 안쪽에 선 추가
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = inside_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(size=9)

            worksheet['A1'].font = Font(size=16, bold=True)
            worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
            worksheet.page_setup.fitToPage = True
            worksheet.page_setup.fitToHeight = 1
            worksheet.page_margins = PageMargins(top=0, bottom=0, left=0, right=0)

            print("엑셀 파일이 생성되었습니다.")

exit()
